#!/usr/bin/python3
# -*- coding: utf-8 -*-

import openpyxl
import sys
import os
import copy
import datetime

class NotMemberException(Exception):
    pass

class WrongBookException(Exception):
    pass

class Hyoka:
    def __init__(self, room, name, laboratory, category, number, shidou, present_cat, present_num, 
                 student, title, tech_point, qa_point, comment):
        self.room = room
        self.name = name
        self.laboratory = laboratory
        self.category = category
        self.number = number
        self.shidou = shidou
        self.present_cat = present_cat
        self.present_num = present_num
        self.student = student
        self.title = title
        self.tech_point = tech_point
        self.qa_point = qa_point
        self.comment = comment

        self.is_student = self._is_student()

    def __str__(self):
        return ("%s %s%d %s %s - 発表=%s 質疑=%s コメント=%s" % 
                (self.shidou, self.present_cat, self.present_num, self.student, self.title,
                 (self.tech_point if self.tech_point else ''),
                 (self.qa_point if self.qa_point else ''),
                 self.comment))
#        return ("(%s) %s %s %s %s-%d - %s %s%d %s 発表=%s 質疑=%s コメント=%s" % 
#                (('学生' if self.is_student else '教員'), 
#                 self.room, self.name, self.laboratory, self.category, self.number,
#                 self.shidou, self.present_cat, self.present_num, self.student, 
#                 (self.tech_point if self.tech_point else ''),
#                 (self.qa_point if self.qa_point else ''),
#                 self.comment))

    def _is_student(self):
        lab = copy.copy(self.laboratory)
        lab.replace('研究室','')
        lab.replace(' ', '')
        lab.replace('　','')
        return (len(lab) > 0)

def get_room(category, number):
    return "%s室(%s)" % (category, ("AM" if number < 200 else "PM"))

def get_hyokasha(category, number, book):
    room = get_room(category, number)
    sheet = book[room]
    
    try:
        name = sheet['H1'].value.replace(' ','').replace('　','')
    except:
        name = "(NO NAME)"
    try:
        laboratory = sheet['H2'].value.replace(' ','').replace('　','')
    except:
        laboratory = "(NO NAME)"

    return (room, name, laboratory, sheet)

def get_hyokas(category, number, book):
    (room, name, laboratory, sheet) = get_hyokasha(category, number, book)

    hyokas = []
    shidou = None
    for row in sheet.iter_rows(min_row=5):        
        if row[0].value:
            shidou = row[0].value

        present_cat = row[1].value
        if not present_cat:
            break
        try:
            present_num = int(row[2].value)
        except:
            # if present_num (ex: 101 of A-101) is not number,
            # this row is not corrent data
            continue

        student = row[3].value
        title = row[4].value

        tech_point = row[5].value
        qa_point = row[6].value
        # check point data
        if not tech_point and not qa_point:
            # if both of tech_point and qa_point are not number value,
            # skip it.
            continue

        comment = row[7].value if row[7].value else ""

        hyoka = Hyoka(room=room, name=name, laboratory=laboratory,
                        category=category, number=number,
                        shidou=shidou, present_cat=present_cat, present_num=present_num,
                        student=student, title=title, 
                        tech_point=tech_point,
                        qa_point=qa_point, 
                        comment=comment)
        #print(hyoka)
        hyokas.append(hyoka)
        
    return hyokas

def check_correct_book(category, number, book):
    room = get_room(category, number)
    try:
        sheet = book[room]
    except:
        raise WrongBookException("a sheet '%s' not found" % room)

    try:
        cell = 'F4'
        if sheet[cell].value != '発表技術':
            raise WrongBookException("the cell '%s' is not '発表技術'" % cell)
    except:
        raise WrongBookException("the cell '%s' is not '発表技術'" % cell)

    try:
        cell = 'G4'
        if sheet[cell].value != '質疑応答':
            raise WrongBookException("the cell '%s' is not '質疑応答'" % cell)
    except:
        raise WrongBookException("the cell '%s' is not '質疑応答'" % cell)

    # in order to came here, the book is correct


def update_shukei(shukei_book, category, number, hyoka_book):
    # check that the book is correct or not
    try:
        check_correct_book(category, number, hyoka_book)
    except WrongBookException as e:
        print("WARN: This book may be wrong %s-%d - %s" % (category, number, e))
        return

    hyokas = get_hyokas(category, number, hyoka_book)
    if len(hyokas) == 0:
        print("WARN: point data is not found - %s-%d" % (category, number))
        return

    (room, hyokasha, laboratory, _) = get_hyokasha(category, number, hyoka_book)
    print("%s %s-%d %s %s" % (room, category, number, hyokasha, laboratory))
    
    for hyoka in hyokas:
        try:
            sheet = shukei_book[hyoka.room]
        except KeyError:
            print("ERROR: sheet not found '%s' in shukei book - %s" % (hyoka.room, hyoka))
            continue

        # AN = 40
        # 102 -> (hyoka.number % 100 - 1) * 2 = 2 -> 39 + 2 = 41 -> AP
        column_num = 40 + (hyoka.number % 100 - 1) * 2
        cell = sheet.cell(row=3, column=column_num)
        if cell.value != hyoka.number:
            # ichiou check suru
            print("ERROR: student number is wrong - number=%d vs %s=%d" %
                    (hyoka.number, cell.coordinate, cell.value if cell.value else -1))
            continue

        # 5 -> 101 or 201
        # 6 -> 102 or 202
        row_num = 5 + (hyoka.present_num % 100 - 1)
        cell = sheet.cell(row=row_num, column=3)
        if not cell.value:
            # empty cell
            continue
        if cell.value != hyoka.present_num:
            # ichiou check suru
            print("ERROR: student number is wrong - present_num=%d vs %s=%d" %
                    (hyoka.present_num, cell.coordinate, (cell.value if cell.value else -1)))
            continue

        tech_point_cell = sheet.cell(row=row_num, column=column_num)
        qa_point_cell = sheet.cell(row=row_num, column=column_num+1)
        tech_point_cell.value = hyoka.tech_point
        qa_point_cell.value = hyoka.qa_point
        print("Updated %s#%s:%s - %s" % 
                (sheet.title, tech_point_cell.coordinate, qa_point_cell.coordinate, hyoka))

if __name__ == '__main__':
    shukei_xls = sys.argv[1]
    basedir = os.path.dirname(shukei_xls)
    shukei_book = openpyxl.load_workbook(shukei_xls)
    for c in ['A', 'B', 'C', 'D']:
        for n in [i for i in range(101, 121)] + [i for i in range(201, 221)]:
            try:
                xls_name = "%s-%d.xlsx" % (c, n)
                xls_path = os.path.join(basedir, xls_name)
                if not os.path.exists(xls_path):
                    continue

                print("Loading %s ... " % xls_path)
                hyoka_book = openpyxl.load_workbook(xls_path, read_only=True, data_only=True)
                update_shukei(shukei_book, c, n, hyoka_book)
            except:
                print("ERROR: %s\n%s" % (sys.exc_info()[0], sys.exc_info[2]))

    filename = os.path.basename(shukei_xls)
    filebase = os.path.splitext(filename)[0]
    fileext = os.path.splitext(filename)[1]
    newpath = os.path.join(basedir, filebase + "_auto_" + datetime.datetime.now().strftime('%Y%m%d-%H%M%S') + fileext)
    shukei_book.save(newpath)
    print("Saved to %s" % newpath)

